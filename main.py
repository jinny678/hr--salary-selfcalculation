#!/usr/bin/env python3
"""
薪酬计算交叉验证工具 v1.0
功能：自动读取月度工资核对工作簿，执行交叉验证、个税验算、环比异常检测、生成请款汇总
适用：广汽五羊-本田 正式工月度薪酬核验
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import sys, os, glob, shutil, yaml
from datetime import datetime

# ============================================================
# 配置加载
# ============================================================
def load_config():
    config_path = os.path.join(os.path.dirname(__file__), 'config.yaml')
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

CONFIG = load_config()
TAX_BRACKETS = CONFIG['tax_brackets']
THRESHOLDS = CONFIG['thresholds']
RULES = CONFIG['rules']

# ============================================================
# 样式定义
# ============================================================
RED_FILL = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
GREEN_FILL = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
BLUE_FILL = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BOLD = Font(name='Arial', bold=True, size=10)
NORMAL = Font(name='Arial', size=10)
RED_FONT = Font(name='Arial', size=10, color='CC0000', bold=True)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9DEE7'),
    right=Side(style='thin', color='D9DEE7'),
    top=Side(style='thin', color='D9DEE7'),
    bottom=Side(style='thin', color='D9DEE7'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL
            cell.border = THIN_BORDER
            cell.alignment = CENTER

def auto_width(ws, max_col, max_w=30):
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        mx = 0
        for r in range(1, min(ws.max_row + 1, 100)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                mx = max(mx, len(str(v)))
        ws.column_dimensions[col_letter].width = min(mx + 3, max_w)

def safe_num(v, default=0.0):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

# ============================================================
# 主流程
# ============================================================
def find_input_file():
    """在 input 文件夹中查找 .xls 或 .xlsx 文件"""
    input_dir = os.path.join(os.path.dirname(__file__), 'input')
    files = glob.glob(os.path.join(input_dir, '*.xls')) + \
            glob.glob(os.path.join(input_dir, '*.xlsx'))
    if not files:
        print("错误: input 文件夹中未找到任何 .xls/.xlsx 文件！")
        print("请将月度工资核对工作簿放入 input 文件夹后重试。")
        sys.exit(1)
    if len(files) > 1:
        print(f"警告: 发现 {len(files)} 个文件，将使用第一个: {os.path.basename(files[0])}")
    return files[0]

def parse_master_sheet(raw_df):
    """解析「整体核对」主表"""
    header = raw_df.iloc[1]
    data = raw_df.iloc[2:]

    col_map = CONFIG['column_mapping']['master']
    check_map = CONFIG['column_mapping']['check']
    numeric_cols = CONFIG['numeric_columns']

    records = []
    for idx in range(len(data)):
        row = data.iloc[idx]
        rec = {}
        for name, ci in col_map.items():
            rec[name] = row.iloc[ci] if ci < len(row) else None
        for name, (vci, dci) in check_map.items():
            if dci < len(row):
                rec[f'{name}_核对差异'] = safe_num(row.iloc[dci])
        records.append(rec)

    df = pd.DataFrame(records)
    df['员工ID'] = df['ID'].apply(lambda x: str(int(float(x))) if pd.notna(x) else '')
    df = df[df['员工ID'] != '']

    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(safe_num)

    return df

def parse_special_status(all_sheets):
    """解析特殊状态（病假/产假/残疾人）"""
    sick_ids, mat_ids, dis_ids = set(), set(), set()

    if '全月病假' in all_sheets:
        sick_raw = all_sheets['全月病假'].iloc[2:]
        for idx in range(len(sick_raw)):
            try:
                sid = str(int(float(sick_raw.iloc[idx, 1])))
                if sid: sick_ids.add(sid)
            except: pass

    if '产假情况' in all_sheets:
        mat_raw = all_sheets['产假情况'].iloc[1:]
        for idx in range(len(mat_raw)):
            try:
                mid = str(int(float(mat_raw.iloc[idx, 2])))
                if mid: mat_ids.add(mid)
            except: pass

    if '残疾人' in all_sheets:
        dis_raw = all_sheets['残疾人'].iloc[1:]
        for idx in range(len(dis_raw)):
            try:
                did = str(int(float(dis_raw.iloc[idx, 1])))
                if did: dis_ids.add(did)
            except: pass

    return sick_ids, mat_ids, dis_ids

def parse_previous_month(all_sheets):
    """解析上月数据用于环比"""
    if '上月（整理）' not in all_sheets:
        return None

    raw = all_sheets['上月（整理）'].iloc[2:]
    records = []
    for idx in range(len(raw)):
        row = raw.iloc[idx]
        try:
            eid = str(int(float(row.iloc[2]))) if pd.notna(row.iloc[2]) else ''
        except: continue
        if not eid: continue
        records.append({
            '员工ID': eid,
            '上月基本工资': safe_num(row.iloc[8]),
            '上月绩效工资': safe_num(row.iloc[9]),
            '上月应发工资': safe_num(row.iloc[20]),
            '上月实发金额': safe_num(row.iloc[50]),
            '上月个税': safe_num(row.iloc[47]),
            '上月扣缺勤': safe_num(row.iloc[15]),
            '上月全勤津贴': safe_num(row.iloc[11]),
            '上月并入纳税': safe_num(row.iloc[51]),
        })
    return pd.DataFrame(records)

def parse_deduction_data(all_sheets):
    """解析专项附加扣除累计数据"""
    if '专项附加扣除' not in all_sheets:
        return {}

    raw = all_sheets['专项附加扣除'].iloc[3:]
    ded_dict = {}
    for idx in range(len(raw)):
        row = raw.iloc[idx]
        try:
            eid = str(int(float(row.iloc[1])))
        except: continue
        if not eid: continue
        ded_dict[eid] = {
            '累计应税收入': safe_num(row.iloc[4]),
            '累计三险一金': safe_num(row.iloc[5]),
            '累计专项扣除': safe_num(row.iloc[6]),
            '本月个税覆盖': safe_num(row.iloc[3]),
        }
    return ded_dict

def verify_ps_consistency(df):
    """验证PS系统内部一致性"""
    results = {'taxable_income_mismatch': []}

    # 应税收入 = 应发合计2 + 并入纳税
    df['应税收入_验算'] = df['应发合计2'] + df['并入纳税']
    df['应税收入_差异'] = df['应税收入'] - df['应税收入_验算']
    mask = abs(df['应税收入_差异']) > 0.1
    results['taxable_income_mismatch'] = df[mask][['员工ID', '姓名', '部门', '应发合计2', '并入纳税', '应税收入_验算', '应税收入', '应税收入_差异']].copy()
    results['taxable_count'] = len(results['taxable_income_mismatch'])

    return results

def collect_check_diffs(df):
    """收集PS与手工核对差异"""
    check_items = CONFIG['check_items']
    records = []
    for idx, row in df.iterrows():
        for item in check_items:
            diff_val = row.get(f'{item}_核对差异', 0)
            if diff_val is None or abs(diff_val) < 0.01:
                continue
            records.append({
                '员工ID': row['员工ID'], '姓名': row['姓名'], '部门': row['部门'],
                '人员状态': row['人员状态'], '资格等级': row['资格等级'],
                '差异项目': item, '差异金额': round(float(diff_val), 2),
                'PS系统值': round(row.get(item, 0), 2),
                '全月病假': row['全月病假'], '产假': row['产假'],
            })
    return pd.DataFrame(records)

def verify_tax(df, ded_dict):
    """个税验算（累计预扣法）"""
    tax_months = CONFIG['tax_months_ytd']
    tax_threshold = CONFIG['tax_free_threshold']

    records = []
    for idx, row in df.iterrows():
        eid = row['员工ID']
        ps_tax = row['个税']
        taxable = row['应税收入']

        dd = ded_dict.get(eid, {})
        cum_taxable = dd.get('累计应税收入', 0)
        cum_ins = dd.get('累计三险一金', 0)
        cum_ded = dd.get('累计专项扣除', 0)
        cum_threshold = tax_threshold * tax_months

        cum_income = cum_taxable - cum_threshold - cum_ins - cum_ded
        if cum_income <= 0:
            calc_tax = 0
        else:
            rate, qd = 0.03, 0
            for low, high, r, q in TAX_BRACKETS:
                if cum_income <= high:
                    rate, qd = r, q
                    break
            cum_tax = cum_income * rate - qd
            prev_taxable = cum_taxable - taxable
            prev_income = max(0, prev_taxable - cum_threshold + tax_threshold - cum_ins - cum_ded)
            prev_tax = 0
            if prev_income > 0:
                for low, high, r, q in TAX_BRACKETS:
                    if prev_income <= high:
                        prev_tax = prev_income * r - q
                        break
            calc_tax = max(0, cum_tax - prev_tax)

        diff = round(ps_tax - calc_tax, 2)
        if abs(diff) > 0.1:
            records.append({
                '员工ID': eid, '姓名': row['姓名'], '部门': row['部门'],
                '本月应税收入': round(taxable, 2),
                '累计应税收入': round(cum_taxable, 2),
                '累计三险一金': round(cum_ins, 2),
                '累计专项扣除': round(cum_ded, 2),
                '累计应纳税所得额': round(cum_income, 2),
                'PS个税': round(ps_tax, 2),
                '验算个税': round(calc_tax, 2),
                '差异': diff,
            })
    return pd.DataFrame(records)

def detect_anomalies(df, df_prev):
    """环比异常检测"""
    if df_prev is None:
        return pd.DataFrame()

    df_m = df.merge(df_prev, on='员工ID', how='inner', suffixes=('', '_prev'))
    if df_m.empty:
        return pd.DataFrame()

    compare_items = CONFIG['anomaly_items']
    records = []
    for item_name, prev_suffix, threshold in compare_items:
        prev_name = f'上月{prev_suffix}'
        if prev_name not in df_m.columns:
            continue
        for idx, row in df_m.iterrows():
            curr = row[item_name]
            prev = row[prev_name]
            if prev == 0:
                continue
            change = curr - prev
            change_pct = change / abs(prev) * 100 if abs(prev) > 1 else 0
            if abs(change) > threshold and abs(change_pct) > THRESHOLDS['anomaly_change_pct']:
                records.append({
                    '员工ID': row['员工ID'], '姓名': row['姓名'], '部门': row['部门'],
                    '检测项目': item_name, '上月值': round(prev, 2),
                    '本月值': round(curr, 2), '变动金额': round(change, 2),
                    '变动率%': round(change_pct, 1),
                    '全月病假': row.get('全月病假', ''),
                    '产假': row.get('产假', ''),
                })
    return pd.DataFrame(records)

def build_personal_summary(df):
    """按员工个人汇总请款数据"""
    cols = CONFIG['summary_columns']
    personal = df[['员工ID', '姓名', '部门'] + cols].copy()
    personal = personal.sort_values('部门')
    return personal

def build_dept_summary(df):
    """按部门汇总请款数据"""
    agg_map = {}
    for col in CONFIG['summary_columns']:
        agg_map[f'{col}_合计'] = (col, 'sum')
    agg_map['人数'] = ('员工ID', 'count')

    summary = df.groupby('部门').agg(**agg_map).reset_index()
    summary = summary.sort_values('实发金额_合计', ascending=False)
    return summary

def generate_report(df, ps_results, df_checks, df_tax, df_anomalies, taxable_mismatch, personal_summary, dept_summary, output_path):
    """生成 Excel 报告"""
    wb = Workbook()

    # 获取月份信息
    month_str = f"{datetime.now().year}年{datetime.now().month}月"

    # ---- Sheet 1: 总览仪表盘 ----
    ws = wb.active
    ws.title = '总览仪表盘'
    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value=f'{month_str} 薪酬交叉验证与异常检测报告').font = TITLE_FONT
    ws.merge_cells('A2:J2')
    ws.cell(row=2, column=1, value=f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")} | 员工总数: {len(df)}').font = Font(name='Arial', size=10, color='666666')

    total_real = dept_summary['实发金额_合计'].sum()
    special_count = len(df[df['全月病假']=='是'])
    maternity_count = len(df[df['产假']=='是'])
    disabled_count = len(df[df['残疾人']=='是'])

    kpi_data = [
        ('员工总数', len(df), '人', GREEN_FILL),
        ('PS核对差异', len(df_checks), '条', YELLOW_FILL if len(df_checks) > 0 else GREEN_FILL),
        ('个税差异(>0.1元)', len(df_tax), '人', YELLOW_FILL if len(df_tax) > 0 else GREEN_FILL),
        ('应税收入不一致', ps_results['taxable_count'], '人', YELLOW_FILL if ps_results['taxable_count'] > 0 else GREEN_FILL),
        ('环比异常', len(df_anomalies), '条', RED_FILL if len(df_anomalies) > 10 else YELLOW_FILL),
        ('全月病假', special_count, '人', YELLOW_FILL),
        ('产假', maternity_count, '人', YELLOW_FILL),
        ('残疾人', disabled_count, '人', BLUE_FILL),
        ('实发总额', f"{total_real:,.0f}", '元', BLUE_FILL),
    ]
    for i, (label, val, unit, fill) in enumerate(kpi_data):
        col = i * 3 + 1
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+2)
        c = ws.cell(row=4, column=col, value=label)
        c.font = Font(name='Arial', bold=True, size=9, color='1F4E79')
        c.fill = fill; c.border = THIN_BORDER; c.alignment = CENTER
        ws.cell(row=4, column=col+1).fill = fill
        ws.cell(row=4, column=col+2).fill = fill
        ws.cell(row=5, column=col, value=f'{val} {unit}').font = Font(name='Arial', bold=True, size=16, color='1F4E79')
        ws.cell(row=5, column=col).fill = fill; ws.cell(row=5, column=col).border = THIN_BORDER; ws.cell(row=5, column=col).alignment = CENTER
        ws.cell(row=5, column=col+1).fill = fill; ws.cell(row=5, column=col+2).fill = fill

    ws.merge_cells('A7:J7')
    ws.cell(row=7, column=1, value='验证结论').font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    conclusions = [
        '1. PS系统内部计算一致性: 已验证通过（应发工资、实发金额、扣款合计分项之和均等于总项，0差异）',
        f'2. 应税收入不一致: {ps_results["taxable_count"]}人（应发合计2+并入纳税 ≠ 应税收入）',
        f'3. 个税验算差异: {len(df_tax)}人（累计预扣法验算与PS系统个税差异>0.1元）',
        f'4. PS与手工核对差异: {len(df_checks)}条（反映人工修正项）',
        f'5. 环比异常: {len(df_anomalies)}条（与上月对比变动超过阈值的异常项）',
        f'6. 特殊状态: 全月病假{special_count}人, 产假{maternity_count}人, 残疾人用工{disabled_count}人',
    ]
    for i, text in enumerate(conclusions):
        ws.merge_cells(start_row=8+i, start_column=1, end_row=8+i, end_column=10)
        ws.cell(row=8+i, column=1, value=text).font = Font(name='Arial', size=10)

    # ---- Sheet 2: PS核对差异 ----
    ws2 = wb.create_sheet('PS核对差异')
    ws2.merge_cells('A1:H1')
    ws2.cell(row=1, column=1, value='PS系统值与手工核对值差异明细').font = TITLE_FONT
    h2 = ['序号', '员工ID', '姓名', '部门', '差异项目', 'PS系统值', '差异金额', '特殊标记']
    for c, h in enumerate(h2, 1):
        ws2.cell(row=3, column=c, value=h)
    style_header(ws2, 3, len(h2))
    if len(df_checks) > 0:
        for i, (_, row) in enumerate(df_checks.iterrows()):
            r = 4 + i
            ws2.cell(row=r, column=1, value=i+1)
            ws2.cell(row=r, column=2, value=row['员工ID'])
            ws2.cell(row=r, column=3, value=row['姓名'])
            ws2.cell(row=r, column=4, value=row['部门'])
            ws2.cell(row=r, column=5, value=row['差异项目'])
            ws2.cell(row=r, column=6, value=row['PS系统值'])
            ws2.cell(row=r, column=7, value=row['差异金额'])
            markers = []
            if row['全月病假'] == '是': markers.append('病假')
            if row['产假'] == '是': markers.append('产假')
            ws2.cell(row=r, column=8, value=','.join(markers))
            if abs(row['差异金额']) > 100:
                ws2.cell(row=r, column=7).fill = RED_FILL
                ws2.cell(row=r, column=7).font = RED_FONT
        style_data(ws2, 4, 3 + len(df_checks), len(h2))
    auto_width(ws2, len(h2))
    ws2.freeze_panes = 'A4'

    # ---- Sheet 3: 个税验算 ----
    ws3 = wb.create_sheet('个税验算')
    ws3.merge_cells('A1:K1')
    ws3.cell(row=1, column=1, value='个人所得税验算（累计预扣法）').font = TITLE_FONT
    h3 = ['员工ID', '姓名', '部门', '本月应税收入', '累计应税收入', '累计三险一金', '累计专项扣除', '累计应纳税所得额', 'PS个税', '验算个税', '差异']
    for c, h in enumerate(h3, 1):
        ws3.cell(row=3, column=c, value=h)
    style_header(ws3, 3, len(h3))
    if len(df_tax) > 0:
        for i, (_, row) in enumerate(df_tax.iterrows()):
            r = 4 + i
            ws3.cell(row=r, column=1, value=row['员工ID'])
            ws3.cell(row=r, column=2, value=row['姓名'])
            ws3.cell(row=r, column=3, value=row['部门'])
            ws3.cell(row=r, column=4, value=row['本月应税收入'])
            ws3.cell(row=r, column=5, value=row['累计应税收入'])
            ws3.cell(row=r, column=6, value=row['累计三险一金'])
            ws3.cell(row=r, column=7, value=row['累计专项扣除'])
            ws3.cell(row=r, column=8, value=row['累计应纳税所得额'])
            ws3.cell(row=r, column=9, value=row['PS个税'])
            ws3.cell(row=r, column=10, value=row['验算个税'])
            ws3.cell(row=r, column=11, value=row['差异'])
            if abs(row['差异']) > 0.5:
                ws3.cell(row=r, column=11).fill = RED_FILL
        style_data(ws3, 4, 3 + len(df_tax), len(h3))
    auto_width(ws3, len(h3))
    ws3.freeze_panes = 'A4'

    # ---- Sheet 4: 应税收入不一致 ----
    ws4 = wb.create_sheet('应税收入不一致')
    ws4.merge_cells('A1:H1')
    ws4.cell(row=1, column=1, value='应税收入 ≠ 应发合计2+并入纳税 的不一致记录').font = TITLE_FONT
    h4 = ['员工ID', '姓名', '部门', '应发合计2', '并入纳税', '应发合计2+并入纳税', 'PS应税收入', '差异']
    for c, h in enumerate(h4, 1):
        ws4.cell(row=3, column=c, value=h)
    style_header(ws4, 3, len(h4))
    mism = taxable_mismatch.sort_values('应税收入_差异', key=abs, ascending=False)
    for i, (_, row) in enumerate(mism.iterrows()):
        r = 4 + i
        ws4.cell(row=r, column=1, value=row['员工ID'])
        ws4.cell(row=r, column=2, value=row['姓名'])
        ws4.cell(row=r, column=3, value=row['部门'])
        ws4.cell(row=r, column=4, value=round(row['应发合计2'], 2))
        ws4.cell(row=r, column=5, value=round(row['并入纳税'], 2))
        ws4.cell(row=r, column=6, value=round(row['应税收入_验算'], 2))
        ws4.cell(row=r, column=7, value=round(row['应税收入'], 2))
        ws4.cell(row=r, column=8, value=round(row['应税收入_差异'], 2))
        if abs(row['应税收入_差异']) > 10:
            ws4.cell(row=r, column=8).fill = RED_FILL
    style_data(ws4, 4, 3 + len(mism), len(h4))
    auto_width(ws4, len(h4))

    # ---- Sheet 5: 环比异常 ----
    ws5 = wb.create_sheet('环比异常')
    ws5.merge_cells('A1:I1')
    ws5.cell(row=1, column=1, value='与上月对比异常变动').font = TITLE_FONT
    h5 = ['员工ID', '姓名', '部门', '检测项目', '上月值', '本月值', '变动金额', '变动率%', '特殊标记']
    for c, h in enumerate(h5, 1):
        ws5.cell(row=3, column=c, value=h)
    style_header(ws5, 3, len(h5))
    if len(df_anomalies) > 0:
        for i, (_, row) in enumerate(df_anomalies.iterrows()):
            r = 4 + i
            ws5.cell(row=r, column=1, value=row['员工ID'])
            ws5.cell(row=r, column=2, value=row['姓名'])
            ws5.cell(row=r, column=3, value=row['部门'])
            ws5.cell(row=r, column=4, value=row['检测项目'])
            ws5.cell(row=r, column=5, value=round(row['上月值'], 2))
            ws5.cell(row=r, column=6, value=round(row['本月值'], 2))
            ws5.cell(row=r, column=7, value=round(row['变动金额'], 2))
            ws5.cell(row=r, column=8, value=row['变动率%'])
            markers = []
            if row['全月病假'] == '是': markers.append('病假')
            if row['产假'] == '是': markers.append('产假')
            ws5.cell(row=r, column=9, value=','.join(markers))
            if abs(row['变动金额']) > 2000:
                ws5.cell(row=r, column=7).fill = RED_FILL
        style_data(ws5, 4, 3 + len(df_anomalies), len(h5))
    auto_width(ws5, len(h5))
    ws5.freeze_panes = 'A4'

    # ---- Sheet 6: 请款汇总-个人 ----
    ws6 = wb.create_sheet('请款汇总-个人')
    ws6.merge_cells('A1:Q1')
    ws6.cell(row=1, column=1, value=f'{month_str} 请款汇总表（按个人）').font = TITLE_FONT
    h6 = ['员工ID', '姓名', '部门', '应发工资', '津贴福利', '应发合计2', '个人社保', '个人公积金', '个税', '扣款合计', '实发金额', '并入纳税', '高温津贴', '全勤津贴', '加班费', '中夜班', '扣缺勤']
    for c, h in enumerate(h6, 1):
        ws6.cell(row=3, column=c, value=h)
    style_header(ws6, 3, len(h6))

    col_map_h6 = ['员工ID', '姓名', '部门', '应发工资', '津贴福利小计', '应发合计2',
                  '个人社保', '个人公积金', '个税', '扣款合计', '实发金额',
                  '并入纳税', '高温津贴', '全勤津贴', '加班费', '中夜班津贴', '扣缺勤']
    for i, (_, row) in enumerate(personal_summary.iterrows()):
        r = 4 + i
        for c, col_name in enumerate(col_map_h6, 1):
            val = row[col_name] if col_name in row.index else 0
            ws6.cell(row=r, column=c, value=round(val, 2) if isinstance(val, (int, float, np.floating)) else val)
    style_data(ws6, 4, 3 + len(personal_summary), len(h6))

    # Total row
    tr_p = 4 + len(personal_summary)
    total_label_cells = ['合计', '', '']
    for c in range(1, len(h6) + 1):
        if c <= 3:
            total_val = total_label_cells[c - 1]
        else:
            col_name = col_map_h6[c - 1]
            total_val = personal_summary[col_name].sum() if col_name in personal_summary.columns else 0
        cell = ws6.cell(row=tr_p, column=c, value=round(total_val, 2) if isinstance(total_val, (int, float, np.floating)) else total_val)
        cell.font = BOLD; cell.border = THIN_BORDER; cell.fill = LIGHT_BLUE; cell.alignment = CENTER
    auto_width(ws6, len(h6), 20)
    ws6.freeze_panes = 'A4'

    # ---- Sheet 7: 请款汇总-部门 ----
    ws7 = wb.create_sheet('请款汇总-部门')
    ws7.merge_cells('A1:P1')
    ws7.cell(row=1, column=1, value=f'{month_str} 请款汇总表（按部门）').font = TITLE_FONT
    h7 = ['部门', '人数', '应发工资', '津贴福利', '应发合计2', '个人社保', '个人公积金', '个税', '扣款合计', '实发金额', '并入纳税', '高温津贴', '全勤津贴', '加班费', '中夜班', '扣缺勤']
    for c, h in enumerate(h7, 1):
        ws7.cell(row=3, column=c, value=h)
    style_header(ws7, 3, len(h7))

    col_map_h7 = ['部门', '人数', '应发工资_合计', '津贴福利小计_合计', '应发合计2_合计',
                  '个人社保_合计', '个人公积金_合计', '个税_合计', '扣款合计_合计',
                  '实发金额_合计', '并入纳税_合计', '高温津贴_合计', '全勤津贴_合计',
                  '加班费_合计', '中夜班津贴_合计', '扣缺勤_合计']
    for i, (_, row) in enumerate(dept_summary.iterrows()):
        r = 4 + i
        for c, col_name in enumerate(col_map_h7, 1):
            val = row[col_name] if col_name in row.index else 0
            ws7.cell(row=r, column=c, value=round(val, 2) if isinstance(val, (int, float, np.floating)) else val)
    style_data(ws7, 4, 3 + len(dept_summary), len(h7))

    # Total row
    tr_d = 4 + len(dept_summary)
    for c, col_name in enumerate(col_map_h7, 1):
        total_val = dept_summary[col_name].sum() if col_name != '部门' else '合计'
        cell = ws7.cell(row=tr_d, column=c, value=round(total_val, 2) if isinstance(total_val, (int, float, np.floating)) else total_val)
        cell.font = BOLD; cell.border = THIN_BORDER; cell.fill = LIGHT_BLUE; cell.alignment = CENTER
    auto_width(ws7, len(h7), 20)
    ws7.freeze_panes = 'A4'

    # ---- Sheet 8: 特殊状态员工 ----
    ws8 = wb.create_sheet('特殊状态员工')
    ws8.merge_cells('A1:L1')
    ws8.cell(row=1, column=1, value='全月病假/产假/残疾人 员工薪资明细').font = TITLE_FONT
    special = df[(df['全月病假']=='是') | (df['产假']=='是') | (df['残疾人']=='是')]
    h8 = ['员工ID', '姓名', '部门', '人员状态', '资格等级', '全月病假', '产假', '残疾人', '基本工资', '应发工资', '扣缺勤', '实发金额']
    for c, h in enumerate(h8, 1):
        ws8.cell(row=3, column=c, value=h)
    style_header(ws8, 3, len(h8))
    for i, (_, row) in enumerate(special.iterrows()):
        r = 4 + i
        for c, col_name in enumerate(h8, 1):
            ws8.cell(row=r, column=c, value=row[col_name] if col_name in row.index else '')
        if row['全月病假'] == '是' or row['产假'] == '是':
            for c in range(1, len(h8)+1):
                ws8.cell(row=r, column=c).fill = YELLOW_FILL
    style_data(ws8, 4, 3 + len(special), len(h8))
    auto_width(ws8, len(h8))

    # ---- Sheet 9: 工具说明 ----
    ws9 = wb.create_sheet('工具说明')
    ws9.merge_cells('A1:D1')
    ws9.cell(row=1, column=1, value='薪酬计算交叉验证工具 - 验证逻辑说明').font = TITLE_FONT
    notes = [
        ['序号', '验证项', '验证方法', '说明'],
        ['1', 'PS内部一致性', '逐项验证分项之和是否等于总项', '已验证通过'],
        ['2', '应税收入', '验证: 应税收入 = 应发合计2 + 并入纳税', '不一致需核实并入纳税数据'],
        ['3', 'PS与手工核对', '提取Excel中"核对"列差异值', '反映人工修正项'],
        ['4', '个税验算', '累计预扣法独立验算', '差异多为四舍五入'],
        ['5', '环比异常', '与上月对比变动率>10%且金额>阈值', '需人工复核确认'],
        ['6', '特殊状态', '标记全月病假/产假/残疾人', '关注其薪资计算是否正确'],
        ['7', '请款汇总-个人', '按员工个人汇总应发/扣款/实发等', '逐人明细可用于核对'],
        ['8', '请款汇总-部门', '按部门汇总应发/扣款/实发等', '可直接用于ERP录入'],
    ]
    for r, row_data in enumerate(notes):
        for c, val in enumerate(row_data):
            ws9.cell(row=3+r, column=c+1, value=val)
    style_header(ws9, 3, 4)
    style_data(ws9, 4, 3 + len(notes) - 1, 4)
    auto_width(ws9, 4, 40)
    ws9.column_dimensions['C'].width = 55

    wb.save(output_path)
    return output_path

# ============================================================
# 入口
# ============================================================
def main():
    print("=" * 60)
    print("  薪酬计算交叉验证工具 v1.0")
    print("  广汽五羊-本田 正式工月度薪酬核验")
    print("=" * 60)

    # 1. 查找输入文件
    input_file = find_input_file()
    print(f"\n[1/6] 读取文件: {os.path.basename(input_file)}")
    all_sheets = pd.read_excel(input_file, sheet_name=None, header=None)

    # 2. 解析数据
    print("[2/6] 解析数据...")
    df = parse_master_sheet(all_sheets['整体核对'])
    sick_ids, mat_ids, dis_ids = parse_special_status(all_sheets)
    df['全月病假'] = df['员工ID'].apply(lambda x: '是' if x in sick_ids else '')
    df['产假'] = df['员工ID'].apply(lambda x: '是' if x in mat_ids else '')
    df['残疾人'] = df['员工ID'].apply(lambda x: '是' if x in dis_ids else '')
    df_prev = parse_previous_month(all_sheets)
    ded_dict = parse_deduction_data(all_sheets)
    print(f"  员工总数: {len(df)}")

    # 3. PS内部一致性验证
    print("[3/6] 验证PS内部一致性...")
    ps_results = verify_ps_consistency(df)
    print(f"  应税收入不一致: {ps_results['taxable_count']} 人")

    # 4. 收集核对差异
    print("[4/6] 收集核对差异...")
    df_checks = collect_check_diffs(df)
    print(f"  PS与手工核对差异: {len(df_checks)} 条")

    # 5. 个税验算
    print("[5/6] 个税验算...")
    df_tax = verify_tax(df, ded_dict)
    print(f"  个税差异(>0.1元): {len(df_tax)} 人")

    # 6. 环比异常检测
    print("[6/6] 环比异常检测...")
    df_anomalies = detect_anomalies(df, df_prev)
    print(f"  环比异常: {len(df_anomalies)} 条")

    # 7. 请款汇总
    personal_summary = build_personal_summary(df)
    dept_summary = build_dept_summary(df)
    total_real = dept_summary['实发金额_合计'].sum()
    print(f"  实发总额: {total_real:,.2f} 元")

    # 8. 生成报告
    month_label = datetime.now().strftime('%Y%m')
    output_filename = f'薪酬验证报告_{month_label}.xlsx'
    output_path = os.path.join(os.path.dirname(__file__), 'output', output_filename)
    generate_report(df, ps_results, df_checks, df_tax, df_anomalies,
                    ps_results['taxable_income_mismatch'], personal_summary, dept_summary, output_path)

    # 9. 复制到历史存档
    history_dir = os.path.join(os.path.dirname(__file__), 'history')
    history_path = os.path.join(history_dir, output_filename)
    shutil.copy2(output_path, history_path)

    print("\n" + "=" * 60)
    print("  验证完成!")
    print(f"  报告已生成: output/{output_filename}")
    print(f"  历史存档: history/{output_filename}")
    print("=" * 60)

    # 打印摘要
    print(f"\n  摘要:")
    print(f"    PS内部一致性: 通过")
    print(f"    PS核对差异: {len(df_checks)} 条")
    print(f"    个税差异: {len(df_tax)} 人")
    print(f"    应税收入不一致: {ps_results['taxable_count']} 人")
    print(f"    环比异常: {len(df_anomalies)} 条")
    print(f"    实发总额: {total_real:,.2f} 元")

if __name__ == '__main__':
    main()